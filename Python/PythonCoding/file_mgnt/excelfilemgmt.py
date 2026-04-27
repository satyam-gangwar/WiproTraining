from openpyxl import Workbook, load_workbook
import os


def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Age"])
    sheet.append(["Satyam", 30])
    sheet.append(["Ram", 25])
    workbook.save(filename)




def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
         print(f"Name: {row[0]}, Age: {row[1]}")

def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"{filename} deleted successfully")
    else:
        print(f"{filename} does not exist")

filename = "data.xlsx"
write_excel(filename)
print("Data read from excel file:")
read_excel(filename)
delete_excel(filename)
